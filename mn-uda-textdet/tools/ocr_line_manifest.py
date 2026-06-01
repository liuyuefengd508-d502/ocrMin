"""Utilities for line-level OCR closed-loop based on rule-based line crops.

This module standardizes:
1) expert Excel <-> line-level manifest (JSONL/CSV)
2) round-aware merge for expert feedback

Design constraints:
- No pandas dependency. Uses openpyxl only.
"""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional

import openpyxl


EXPECTED_EXCEL_COLUMNS: tuple[str, ...] = (
    "页面ID",
    "行ID",
    "阅读顺序",
    "bbox_x1",
    "bbox_y1",
    "bbox_x2",
    "bbox_y2",
    "行图像相对路径",
    "转写文本",
    "难以辨认",
    "专家备注",
)


@dataclass(frozen=True)
class LineRecord:
    page_id: str
    line_id: str
    reading_order: int
    bbox: tuple[int, int, int, int]  # x1,y1,x2,y2
    rel_img_path: str
    transcription: str
    illegible: str
    expert_note: str
    round_id: int = 0
    budget_tag: str = ""
    priority_score: float = 0.0

    def to_dict(self) -> dict[str, Any]:
        return {
            "page_id": self.page_id,
            "line_id": self.line_id,
            "reading_order": self.reading_order,
            "bbox": list(self.bbox),
            "rel_img_path": self.rel_img_path,
            "transcription": self.transcription,
            "illegible": self.illegible,
            "expert_note": self.expert_note,
            "round": self.round_id,
            "budget_tag": self.budget_tag,
            "priority_score": self.priority_score,
        }


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v)


def _as_int(v: Any, default: int = 0) -> int:
    if v is None or str(v).strip() == "":
        return default
    try:
        return int(float(v))
    except Exception:
        return default


def _as_float(v: Any, default: float = 0.0) -> float:
    if v is None or str(v).strip() == "":
        return default
    try:
        return float(v)
    except Exception:
        return default


def read_expert_xlsx(path: Path, sheet: str = "all_lines") -> list[LineRecord]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"sheet not found: {sheet}, available={wb.sheetnames}")
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    if header is None:
        raise ValueError("empty excel sheet")
    col_to_i = {str(k).strip(): i for i, k in enumerate(header) if k is not None}

    missing = [c for c in EXPECTED_EXCEL_COLUMNS if c not in col_to_i]
    if missing:
        raise ValueError(f"missing columns in excel: {missing}")

    out: list[LineRecord] = []
    for r in rows:
        if r is None:
            continue
        page_id = _as_str(r[col_to_i["页面ID"]]).strip()
        line_id = _as_str(r[col_to_i["行ID"]]).strip()
        if not page_id or not line_id:
            continue
        rec = LineRecord(
            page_id=page_id,
            line_id=line_id,
            reading_order=_as_int(r[col_to_i["阅读顺序"]], 0),
            bbox=(
                _as_int(r[col_to_i["bbox_x1"]]),
                _as_int(r[col_to_i["bbox_y1"]]),
                _as_int(r[col_to_i["bbox_x2"]]),
                _as_int(r[col_to_i["bbox_y2"]]),
            ),
            rel_img_path=_as_str(r[col_to_i["行图像相对路径"]]).strip(),
            transcription=_as_str(r[col_to_i["转写文本"]]),
            illegible=_as_str(r[col_to_i["难以辨认"]]),
            expert_note=_as_str(r[col_to_i["专家备注"]]),
        )
        out.append(rec)
    return out


def write_manifest_jsonl(records: Iterable[LineRecord], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        for rec in records:
            f.write(json.dumps(rec.to_dict(), ensure_ascii=False) + "\n")


def write_manifest_csv(records: Iterable[LineRecord], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(LineRecord.__dataclass_fields__.keys())
    # Keep stable and explicit order for CSV consumers.
    fieldnames = [
        "page_id",
        "line_id",
        "reading_order",
        "bbox",
        "rel_img_path",
        "transcription",
        "illegible",
        "expert_note",
        "round_id",
        "budget_tag",
        "priority_score",
    ]
    with out_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for rec in records:
            d = rec.to_dict()
            # Map JSON keys to CSV keys.
            w.writerow(
                {
                    "page_id": d["page_id"],
                    "line_id": d["line_id"],
                    "reading_order": d["reading_order"],
                    "bbox": json.dumps(d["bbox"], ensure_ascii=False),
                    "rel_img_path": d["rel_img_path"],
                    "transcription": d["transcription"],
                    "illegible": d["illegible"],
                    "expert_note": d["expert_note"],
                    "round_id": d["round"],
                    "budget_tag": d["budget_tag"],
                    "priority_score": d["priority_score"],
                }
            )


def _load_jsonl_index(path: Path) -> dict[str, dict[str, Any]]:
    idx: dict[str, dict[str, Any]] = {}
    if not path.exists():
        return idx
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        obj = json.loads(line)
        line_id = str(obj.get("line_id", "")).strip()
        if not line_id:
            continue
        idx[line_id] = obj
    return idx


def merge_expert_into_manifest(
    base_manifest_jsonl: Path,
    expert_xlsx: Path,
    out_manifest_jsonl: Path,
    round_id: int,
    sheet: str = "all_lines",
) -> dict[str, Any]:
    """Merge expert transcription fields into an existing manifest.

    Merge policy (by line_id):
    - Always keep structural fields from base manifest (page_id, bbox, rel_img_path, reading_order).
    - Overwrite `transcription/illegible/expert_note` with non-empty values from expert_xlsx.
    - Set record round to `round_id` iff transcription is non-empty after merge.
    """
    base_idx = _load_jsonl_index(base_manifest_jsonl)
    expert_recs = read_expert_xlsx(expert_xlsx, sheet=sheet)
    updated, filled = 0, 0

    for r in expert_recs:
        cur = base_idx.get(r.line_id)
        if cur is None:
            # If missing, create a minimal record from expert (still useful).
            cur = r.to_dict()
            base_idx[r.line_id] = cur
        before = (str(cur.get("transcription", "")) or "").strip()
        new_t = (r.transcription or "").strip()
        new_il = (r.illegible or "").strip()
        new_note = (r.expert_note or "").strip()

        if new_t:
            cur["transcription"] = r.transcription
            cur["round"] = int(round_id)
        if new_il:
            cur["illegible"] = r.illegible
        if new_note:
            cur["expert_note"] = r.expert_note

        after = (str(cur.get("transcription", "")) or "").strip()
        if after != before:
            updated += 1
        if not before and after:
            filled += 1

    # Write merged manifest deterministically sorted by (page_id, reading_order, line_id).
    def _sort_key(obj: dict[str, Any]) -> tuple:
        return (
            str(obj.get("page_id", "")),
            int(obj.get("reading_order", 0) or 0),
            str(obj.get("line_id", "")),
        )

    out_manifest_jsonl.parent.mkdir(parents=True, exist_ok=True)
    with out_manifest_jsonl.open("w", encoding="utf-8") as f:
        for obj in sorted(base_idx.values(), key=_sort_key):
            f.write(json.dumps(obj, ensure_ascii=False) + "\n")

    return {
        "base_manifest": str(base_manifest_jsonl),
        "expert_xlsx": str(expert_xlsx),
        "out_manifest": str(out_manifest_jsonl),
        "round_id": int(round_id),
        "updated": int(updated),
        "filled": int(filled),
        "total": int(len(base_idx)),
    }

